from requests_type import my_requests
from openpyxl import load_workbook
import json

a01 = load_workbook('testcase01.xlsx')  #获取excel文件
a02 = a01['a03']                        #获取文件里的工作表
rows = a02.max_row                      #获取最大行数
for rows_for in range(2,rows+1):
    #print('当前行号：',index)
    type = a02.cell(rows_for,4).value   #循环获取每一行的各列数据
    url = a02.cell(rows_for,5).value
    data = a02.cell(rows_for,6).value
    headers = a02.cell(rows_for,7).value
    expected_res = a02.cell(rows_for,8).value
                                            #将字符串str转成字典格式
    headers_dict =json.loads(headers)
    # data_dict = json.loads(data)
    data_dict = eval(data)
    # headers_ev =eval(headers)
    res = my_requests(type,url,data_dict,headers_dict)#要求字符类型为字典
    print('响应结果为：',res)
                                                #将响应结果写入excel中
    a02.cell(rows_for,9).value = res[1]
    if res[1] == expected_res:
        print(True)
                                                #将结果写入excel
        a02.cell(rows_for,10).value = '通过'
    # if res[1]['success'] == False:
    #     a02.cell(index,10).value = True
    else:
        print(False)
        a02.cell(rows_for,10).value = '不通过'

#保存修改的数据,需要先关闭excel
a01.save('copy_testcase01.xlsx')