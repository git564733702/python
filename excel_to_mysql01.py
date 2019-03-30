from openpyxl import load_workbook

#处理int数据
#处理var数据

a01 = load_workbook(
    'C:/Users/56473/Desktop/Notepad++/工作簿1.xlsx')
a02 = a01['sh1']
rows = a02.max_row
for index in range(1,rows+1):
    #A列
    b1 = a02.cell(index,1).value
    if isinstance(b1,str):
        b1 = "'" + b1 + "'"
    else:
        b1 = str(b1)
    # B列
    b2 = a02.cell(index,2).value
    if isinstance(b2,str):
        b2 = "'"+b2+"'"
    else:
        b2 = str(b2)
    # C列
    b3 = a02.cell(index,3).value
    if isinstance(b3,str):
        b3 = "'" + b3 + "'"
    else:
        b3 = str(b3)
    # D列
    b4 = a02.cell(index,4).value
    if isinstance(b3,str):
        b4 = "'"+b4+"'"
    else:
        b3 = str(b3)
    print("("+b1+","+b2+","+b3+","+b4+"),")
    #print("("+a1+",'"+b1+"',"+c1+"),")
    #print("("+a1+","+b1+","+c1+"),")
