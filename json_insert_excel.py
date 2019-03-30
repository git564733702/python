from openpyxl import load_workbook
a007 = load_workbook('ningmeng/testcase01.xlsx') #获取excel文件
a02 = a007['a03']                           #获取文件中的工作表

data ={"username":"user","fullname":"user01",
 "email":"123456@qq.com","password1":"123456",
 "password2":"123456","Submit":"Create+account"}
url = 'http://127.0.0.1:8080/jenkins/securityRealm/createAccount'
type = 'post'
headers ={
  "Connection": "keep-alive",
  "Cache-Control": "max-age=0",
  "Upgrade-Insecure-Requests": "1",
  "Content-Type": "application/x-www-form-urlencoded",
  "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36",
  "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3",
  "Accept-Encoding": "gzip, deflate, br",
  "Accept-Language": "zh-CN,zh;q=0.9"
}

i_sta = 8       #第一个账号的尾号
i_count = 10     #账号总数
i = i_sta-1
for index in range(2, i_count+2):
    i +=1
    data["username"]='user'+str(i)      #指定要修改的values，对应的key
    a02.cell(index,4).value = str(type)
    a02.cell(index, 5).value = str(url)
    a02.cell(index, 6).value = str(data)
    # a02.cell(index, 7).value = str(headers)
    # headers的数据导入会变成单引号。请求时会无法转换成json

# 保存修改的数据,需要先关闭excel
a007.save('ningmeng/testcase01.xlsx')